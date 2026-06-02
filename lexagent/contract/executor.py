"""
Playbook executor — runs a declarative detect→grade loop against contract text.

Each position in the PlaybookSpec is evaluated independently:
  1. Detect: ask the LLM whether the clause is present and what it says
  2. Grade: compare against our_position and assign a severity

WHY sequential not parallel:
  Indian commercial contracts tend to be 20–60 positions; a sequential loop
  is easier to checkpoint, cheaper to resume on failure, and avoids rate limits.

WHY xlsx export:
  Lawyers share review results with clients as spreadsheets, not JSON.
"""
from __future__ import annotations

import asyncio
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import litellm

from lexagent.contract.models import (
    PlaybookExecution,
    PlaybookSpec,
    PositionResult,
)


class PlaybookExecutor:
    """
    Runs a PlaybookSpec against a contract document path.
    Returns a completed PlaybookExecution with per-position results.
    """

    def __init__(self, model: str = "anthropic/claude-sonnet-4-6") -> None:
        self._model = model

    async def run(
        self,
        spec: PlaybookSpec,
        document_path: str,
        matter_id: Optional[str] = None,
    ) -> PlaybookExecution:
        """
        Sequential detect→grade loop over all positions in the playbook.
        Returns a PlaybookExecution with status='completed' or 'failed'.
        """
        execution = PlaybookExecution(
            playbook_id=spec.id,
            matter_id=matter_id,
            document_path=document_path,
            status="running",
        )

        try:
            loop = asyncio.get_event_loop()
            contract_text = await loop.run_in_executor(
                None, _extract_text, document_path
            )

            results: list[PositionResult] = []
            for position in spec.positions:
                result = await self._evaluate_position(position, contract_text)
                results.append(result)

            execution.results = results
            execution.overall_risk = _compute_overall_risk(results)
            execution.status = "completed"
            execution.completed_at = datetime.now(tz=timezone.utc).isoformat()

        except Exception as exc:
            execution.status = "failed"
            execution.error = str(exc)
            execution.completed_at = datetime.now(tz=timezone.utc).isoformat()

        return execution

    async def _evaluate_position(self, position, contract_text: str) -> PositionResult:
        """Ask the LLM to detect and grade one position."""
        system = (
            "You are a contract review assistant. Analyse the given contract text "
            "for the specified clause and return ONLY valid JSON with these exact keys:\n"
            '  "detected": bool,\n'
            '  "excerpt": string or null (the relevant clause text, max 300 chars),\n'
            '  "deviation": string or null (how the contract deviates from our position, null if aligned),\n'
            '  "severity": "ok" | "minor" | "major" | "critical"\n'
            "severity is ok when aligned, minor for stylistic gaps, "
            "major for substantive gaps, critical for unacceptable terms."
        )
        user = (
            f"Clause: {position.clause}\n"
            f"Our position: {position.our_position}\n\n"
            f"Contract text:\n---\n{contract_text[:60_000]}\n---\n\n"
            "Return JSON only, no markdown fences."
        )

        try:
            response = await litellm.acompletion(
                model=self._model,
                messages=[
                    {"role": "system", "content": system},
                    {"role": "user", "content": user},
                ],
                response_format={"type": "json_object"},
                request_timeout=60,
            )
            import json
            raw = json.loads(response.choices[0].message.content or "{}")
            return PositionResult(
                clause=position.clause,
                our_position=position.our_position,
                detected=bool(raw.get("detected", False)),
                excerpt=raw.get("excerpt"),
                deviation=raw.get("deviation"),
                severity=raw.get("severity", "ok"),
            )
        except Exception as exc:
            return PositionResult(
                clause=position.clause,
                our_position=position.our_position,
                detected=False,
                deviation=f"Evaluation error: {exc}",
                severity="major",
            )

    def export_xlsx(self, execution: PlaybookExecution, output_path: str) -> str:
        """
        Export a PlaybookExecution to an xlsx spreadsheet.
        Returns the output path. Requires openpyxl (already in deps).

        WHY xlsx: Lawyers share contract reviews as spreadsheets with clients.
        A single sheet with one row per position is the standard format.
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Playbook Review"

        headers = ["Clause", "Our Position", "Detected", "Severity", "Deviation", "Excerpt"]
        _SEVERITY_COLOURS = {
            "ok": "C6EFCE",       # green
            "minor": "FFEB9C",    # yellow
            "major": "FFC7CE",    # light red
            "critical": "FF0000", # red
        }

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)

        for row_idx, result in enumerate(execution.results, 2):
            ws.cell(row=row_idx, column=1, value=result.clause)
            ws.cell(row=row_idx, column=2, value=result.our_position)
            ws.cell(row=row_idx, column=3, value="Yes" if result.detected else "No")
            sev_cell = ws.cell(row=row_idx, column=4, value=result.severity.upper())
            colour = _SEVERITY_COLOURS.get(result.severity, "FFFFFF")
            sev_cell.fill = PatternFill("solid", fgColor=colour)
            ws.cell(row=row_idx, column=5, value=result.deviation or "")
            ws.cell(row=row_idx, column=6, value=result.excerpt or "")

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["E"].width = 50
        ws.column_dimensions["F"].width = 50

        wb.save(output_path)
        return output_path


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _extract_text(document_path: str) -> str:
    """Synchronous PDF/text extraction — called via run_in_executor."""
    path = Path(document_path)
    if path.suffix.lower() == ".pdf":
        import pdfplumber
        parts: list[str] = []
        with pdfplumber.open(str(path)) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    parts.append(t)
        return "\n\n".join(parts)
    return path.read_text(encoding="utf-8")


def _compute_overall_risk(results: list[PositionResult]) -> str:
    if any(r.severity == "critical" for r in results):
        return "HIGH"
    if any(r.severity == "major" for r in results):
        return "HIGH"
    if any(r.severity == "minor" for r in results):
        return "MEDIUM"
    return "LOW"
