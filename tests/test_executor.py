"""Tests for lexagent/contract/executor.py — PlaybookExecutor."""
from __future__ import annotations

import json
from unittest.mock import AsyncMock, MagicMock, patch

import pytest

from lexagent.contract.executor import PlaybookExecutor, _compute_overall_risk
from lexagent.contract.models import PlaybookExecution, PlaybookPosition, PlaybookSpec, PositionResult


def _make_spec() -> PlaybookSpec:
    return PlaybookSpec(
        id="nda",
        name="NDA",
        contract_type="nda",
        positions=[
            PlaybookPosition(clause="Governing law", our_position="Delhi courts"),
            PlaybookPosition(clause="Confidentiality period", our_position="3 years"),
        ],
    )


def _litellm_response(content: dict) -> MagicMock:
    msg = MagicMock()
    msg.content = json.dumps(content)
    choice = MagicMock()
    choice.message = msg
    resp = MagicMock()
    resp.choices = [choice]
    return resp


# ---------------------------------------------------------------------------
# _compute_overall_risk()
# ---------------------------------------------------------------------------

def test_compute_risk_all_ok():
    results = [PositionResult(clause="A", our_position="X", detected=True, severity="ok")]
    assert _compute_overall_risk(results) == "LOW"


def test_compute_risk_minor_gives_medium():
    results = [PositionResult(clause="A", our_position="X", detected=True, severity="minor")]
    assert _compute_overall_risk(results) == "MEDIUM"


def test_compute_risk_major_gives_high():
    results = [PositionResult(clause="A", our_position="X", detected=True, severity="major")]
    assert _compute_overall_risk(results) == "HIGH"


def test_compute_risk_critical_gives_high():
    results = [PositionResult(clause="A", our_position="X", detected=True, severity="critical")]
    assert _compute_overall_risk(results) == "HIGH"


def test_compute_risk_empty_gives_low():
    assert _compute_overall_risk([]) == "LOW"


# ---------------------------------------------------------------------------
# PlaybookExecutor.run()
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_executor_run_success():
    """Happy path: two positions, LLM returns clean JSON for each."""
    executor = PlaybookExecutor(model="anthropic/claude-sonnet-4-6")
    spec = _make_spec()

    ok_response = _litellm_response({
        "detected": True, "excerpt": "Delhi courts", "deviation": None, "severity": "ok"
    })

    with (
        patch("lexagent.contract.executor._extract_text", return_value="Sample contract text"),
        patch("litellm.acompletion", new=AsyncMock(return_value=ok_response)),
    ):
        execution = await executor.run(spec, "/tmp/test.pdf", matter_id="matter_001")

    assert execution.status == "completed"
    assert len(execution.results) == 2
    assert execution.results[0].severity == "ok"
    assert execution.overall_risk == "LOW"
    assert execution.completed_at is not None


@pytest.mark.asyncio
async def test_executor_run_with_deviation():
    """One position deviates → overall risk should be elevated."""
    executor = PlaybookExecutor(model="anthropic/claude-sonnet-4-6")
    spec = _make_spec()

    responses = [
        _litellm_response({"detected": True, "excerpt": "Mumbai courts", "deviation": "Differs", "severity": "major"}),
        _litellm_response({"detected": True, "excerpt": "2 years", "deviation": None, "severity": "ok"}),
    ]

    with (
        patch("lexagent.contract.executor._extract_text", return_value="Contract text here"),
        patch("litellm.acompletion", new=AsyncMock(side_effect=responses)),
    ):
        execution = await executor.run(spec, "/tmp/test.pdf")

    assert execution.status == "completed"
    assert execution.overall_risk == "HIGH"
    deviating = [r for r in execution.results if r.deviation]
    assert len(deviating) == 1


@pytest.mark.asyncio
async def test_executor_run_handles_llm_failure():
    """LLM errors produce a 'major' severity PositionResult, not a crash."""
    executor = PlaybookExecutor()
    spec = _make_spec()

    with (
        patch("lexagent.contract.executor._extract_text", return_value="text"),
        patch("litellm.acompletion", new=AsyncMock(side_effect=Exception("LLM error"))),
    ):
        execution = await executor.run(spec, "/tmp/test.pdf")

    assert execution.status == "completed"
    assert all(r.severity == "major" for r in execution.results)


@pytest.mark.asyncio
async def test_executor_run_fails_on_extraction_error():
    """PDF extraction failure → execution.status='failed'."""
    executor = PlaybookExecutor()
    spec = _make_spec()

    with patch("lexagent.contract.executor._extract_text", side_effect=Exception("No PDF")):
        execution = await executor.run(spec, "/tmp/missing.pdf")

    assert execution.status == "failed"
    assert execution.error is not None


# ---------------------------------------------------------------------------
# PlaybookExecutor.export_xlsx()
# ---------------------------------------------------------------------------

@pytest.mark.skipif(
    __import__("importlib.util", fromlist=["find_spec"]).find_spec("openpyxl") is None,
    reason="openpyxl not installed",
)
def test_export_xlsx_creates_file(tmp_path):
    executor = PlaybookExecutor()
    execution = PlaybookExecution(
        playbook_id="nda",
        document_path="/tmp/x.pdf",
        status="completed",
        overall_risk="MEDIUM",
        results=[
            PositionResult(clause="Governing law", our_position="Delhi courts", detected=True, severity="ok"),
            PositionResult(
                clause="Confidentiality period",
                our_position="3 years",
                detected=True,
                deviation="Contract says 1 year",
                severity="minor",
                excerpt="1 year from execution date",
            ),
        ],
    )

    output = str(tmp_path / "review.xlsx")
    result_path = executor.export_xlsx(execution, output)
    assert result_path == output

    import openpyxl
    wb = openpyxl.load_workbook(output)
    ws = wb.active
    # Header row + 2 data rows
    assert ws.max_row == 3
    # Verify first data row clause name
    assert ws.cell(row=2, column=1).value == "Governing law"
